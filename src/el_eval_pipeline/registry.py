from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .common import is_blankish, normalize_space


def load_domain_skill_registry(path: Path) -> dict[str, dict[str, Any]]:
    if not path.exists():
        return {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_space(cell.value) for cell in sheet[2]]
    rows: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        record = {headers[index]: normalize_space(value) for index, value in enumerate(row) if index < len(headers)}
        name = record.get("技能名", "")
        if is_blankish(name):
            continue
        rows[name] = {
            "name": name,
            "type": record.get("技能类型", ""),
            "purpose": record.get("用途", ""),
            "triggers": record.get("触发提示词/场景", ""),
            "typical_output": record.get("典型输出", ""),
            "source": str(path),
        }
    workbook.close()
    return rows


def load_runtime_skill_registry(sessions_json: Path) -> dict[str, dict[str, Any]]:
    if not sessions_json.exists():
        return {}
    data = json.loads(sessions_json.read_text(encoding="utf-8"))
    skills: dict[str, dict[str, Any]] = {}
    for session in data.values():
        snapshot = session.get("skillsSnapshot", {})
        for skill in snapshot.get("resolvedSkills", []):
            name = skill.get("name")
            if name:
                skills[name] = {
                    "name": name,
                    "description": skill.get("description", ""),
                    "file_path": skill.get("filePath", ""),
                    "source": "sessions.json",
                }
        for skill in snapshot.get("skills", []):
            name = skill.get("name")
            if name and name not in skills:
                skills[name] = {"name": name, "source": "sessions.json"}
    return skills

