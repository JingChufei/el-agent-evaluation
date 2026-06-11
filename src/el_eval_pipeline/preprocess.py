from __future__ import annotations

import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .attachments import build_attachment_manifest, resolve_attachment_cell
from .common import is_blankish, normalize_space, split_multiline, write_csv, write_json, write_jsonl
from .d3 import d3_priority, is_d3_candidate
from .paths import as_posix_relative
from .registry import load_domain_skill_registry, load_runtime_skill_registry


def _truthy_yes(value: Any) -> bool:
    return normalize_space(value) == "是"


def _truthy_no(value: Any) -> bool:
    return normalize_space(value) == "否"


def _cell_map(headers: list[str], row_values: tuple[Any, ...]) -> dict[str, str]:
    return {
        headers[index]: "" if value is None else str(value).strip()
        for index, value in enumerate(row_values)
        if index < len(headers)
    }


def _parse_required_files(value: str) -> list[dict[str, Any]]:
    if is_blankish(value):
        return []
    return [
        {
            "path": part,
            "must_exist": True,
            "source": "file_state",
        }
        for part in split_multiline(value)
        if not is_blankish(part)
    ]


def _parse_expected_answer(value: str) -> dict[str, Any] | None:
    if is_blankish(value):
        return None
    assertions: list[dict[str, Any]] = []
    lines = split_multiline(value)
    labeled_value_pattern = re.compile(r"(?P<label>.+?)\s*[:：=]\s*(?P<rest>.+)")
    number_pattern = re.compile(r"(?P<number>[-+]?\d+(?:\.\d+)?)\s*(?P<unit>%|cd/A|nm|h|V|cd/m²|cd/m2)?")
    for line in lines:
        labeled_value = labeled_value_pattern.match(line.rstrip(",;，；"))
        if labeled_value:
            label = labeled_value.group("label").strip().strip('"').strip("'").strip()
            matches = list(number_pattern.finditer(labeled_value.group("rest")))
            for index, match in enumerate(matches):
                assertions.append(
                    {
                        "type": "numeric_contains",
                        "label": label if len(matches) == 1 else f"{label}[{index}]",
                        "value": float(match.group("number")),
                        "unit": match.group("unit") or "",
                        "tolerance": {"abs": 0.01, "rel": 0.02},
                        "raw": line,
                    }
                )
            if not matches:
                assertions.append({"type": "text_contains", "value": line, "raw": line})
        else:
            assertions.append({"type": "text_contains", "value": line, "raw": line})
    return {"type": "multi_assertion", "raw": value, "assertions": assertions}


def _parse_gold_chain(skill_value: str, tool_value: str) -> dict[str, Any] | None:
    skills = [part for part in split_multiline(skill_value) if not is_blankish(part)]
    tools = [part for part in split_multiline(tool_value) if not is_blankish(part)]
    if not skills and not tools:
        return None
    stages: list[dict[str, Any]] = []
    if skills:
        stages.append({"type": "unordered", "steps": skills, "step_type": "skill"})
    if tools:
        stages.append({"type": "unordered", "steps": tools, "step_type": "tool"})
    return {"stages": stages, "forbidden": [], "allow_extra": True, "recovery_window": 3}


def _target_state(row: dict[str, str]) -> dict[str, Any] | None:
    required_files = _parse_required_files(row.get("文件状态", ""))
    if not required_files:
        return None
    return {
        "required_files": required_files,
        "required_db_states": [],
        "required_tool_artifacts": [],
        "forbidden_states": [],
    }


def load_cases_from_excel(
    excel_path: Path,
    attachments_root: Path,
    *,
    case_prefix: str = "EL260529F",
    repo_root: Path | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    repo_root = repo_root or Path.cwd()
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_space(cell.value) for cell in sheet[2]]
    manifest = build_attachment_manifest(attachments_root, repo_root=repo_root)
    cases: list[dict[str, Any]] = []
    attachment_errors: list[dict[str, Any]] = []

    case_index = 0
    task_type = ""
    for excel_row, row_values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        row = _cell_map(headers, row_values)
        if not any(value for value in row.values()):
            continue
        case_index += 1
        if not is_blankish(row.get("任务类型", "")):
            task_type = row.get("任务类型", "")
        attachments, errors = resolve_attachment_cell(row.get("输入（图片/文件）", ""), manifest)
        case_id = f"{case_prefix}-{case_index:04d}"
        for error in errors:
            attachment_errors.append({"case_id": case_id, "source_row": excel_row, **error})

        d2_match_label = row.get("是否可明确匹配的", "")
        d2_enabled = _truthy_yes(d2_match_label)
        d3_candidate = _truthy_no(d2_match_label)
        expected_answer = _parse_expected_answer(row.get("若是，答案是什么", "")) if d2_enabled else None
        target_state = _target_state(row)
        gold_chain = _parse_gold_chain(row.get("Skill", ""), row.get("Tool", ""))

        annotation_status: list[str] = []
        if d2_enabled and expected_answer is None:
            annotation_status.append("d2_expected_answer_missing")
        if not d2_enabled and not d3_candidate:
            annotation_status.append("d2_match_label_missing_or_unknown")
        if d3_candidate and is_blankish(row.get("参考答案", "")):
            annotation_status.append("d3_reference_answer_missing")
        if errors:
            annotation_status.append("attachment_resolution_error")

        evaluable_dimensions = ["D5", "D8"]
        if target_state:
            evaluable_dimensions.append("D1")
        if d2_enabled:
            evaluable_dimensions.append("D2")
        if d3_candidate:
            evaluable_dimensions.append("D3")
        if gold_chain:
            evaluable_dimensions.append("D4")

        cases.append(
            {
                "case_id": case_id,
                "source_file": as_posix_relative(excel_path, repo_root),
                "source_sheet": sheet.title,
                "source_row": excel_row,
                "task_type": task_type or row.get("任务类型", ""),
                "requires_skill": _truthy_yes(row.get("需要skill", "")),
                "skill_name": row.get("关于skill", ""),
                "skill_description": row.get("skill描述", ""),
                "user_query": row.get("输入（文本）", ""),
                "input_file_cell": row.get("输入（图片/文件）", ""),
                "attachments": attachments,
                "attachment_errors": errors,
                "reference_answer": row.get("参考答案", ""),
                "answer_image": row.get("答案附图", ""),
                "target_state": target_state,
                "gold_chain": gold_chain,
                "d2_enabled": d2_enabled,
                "expected_answer": expected_answer,
                "d3_candidate": d3_candidate,
                "d3_priority": "not_applicable" if not d3_candidate else d3_priority({"target_state": target_state}),
                "raw_annotations": {
                    "文件状态": row.get("文件状态", ""),
                    "数据库状态": row.get("数据库状态", ""),
                    "工具产物": row.get("工具产物", ""),
                    "禁止状态": row.get("禁止状态", ""),
                    "Skill": row.get("Skill", ""),
                    "Tool": row.get("Tool", ""),
                    "是否可明确匹配的": row.get("是否可明确匹配的", ""),
                    "若是，答案是什么": row.get("若是，答案是什么", ""),
                },
                "annotation_status": annotation_status,
                "evaluable_dimensions": sorted(evaluable_dimensions),
            }
        )
    workbook.close()
    return cases, manifest, attachment_errors


def build_quality_report(
    cases: list[dict[str, Any]],
    manifest: list[dict[str, Any]],
    attachment_errors: list[dict[str, Any]],
    *,
    domain_skills: dict[str, dict[str, Any]] | None = None,
    runtime_skills: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    domain_skills = domain_skills or {}
    runtime_skills = runtime_skills or {}
    known_skills = set(domain_skills) | set(runtime_skills)
    skill_counter = Counter(case.get("skill_name", "") for case in cases)
    d2_missing = [
        {"case_id": case["case_id"], "source_row": case["source_row"], "skill": case["skill_name"], "user_query": case["user_query"]}
        for case in cases
        if case.get("d2_enabled") and case.get("expected_answer") is None
    ]
    d3_candidates = [
        {"case_id": case["case_id"], "source_row": case["source_row"], "skill": case["skill_name"], "user_query": case["user_query"], "d3_priority": case.get("d3_priority")}
        for case in cases
        if is_d3_candidate(case)
    ]
    d3_missing_reference = [
        {"case_id": case["case_id"], "source_row": case["source_row"], "skill": case["skill_name"], "user_query": case["user_query"]}
        for case in cases
        if is_d3_candidate(case) and is_blankish(case.get("reference_answer", ""))
    ]
    unknown_skills = sorted(
        {
            case["skill_name"]
            for case in cases
            if case.get("skill_name") and case.get("requires_skill") and case["skill_name"] not in known_skills
        }
    )
    coverage = {
        "D1": sum(1 for case in cases if case.get("target_state")),
        "D2": sum(1 for case in cases if case.get("d2_enabled")),
        "D2_runnable": sum(1 for case in cases if case.get("d2_enabled") and case.get("expected_answer")),
        "D3": len(d3_candidates),
        "D3_rubric_synthesis_ready": len(d3_candidates) - len(d3_missing_reference),
        "D4": sum(1 for case in cases if case.get("gold_chain")),
        "D5": len(cases),
        "D8": len(cases),
    }
    return {
        "case_count": len(cases),
        "attachment_count": len(manifest),
        "explicit_attachment_case_count": sum(1 for case in cases if not is_blankish(case.get("input_file_cell", ""))),
        "attachment_error_count": len(attachment_errors),
        "coverage": coverage,
        "d2_missing_answer_count": len(d2_missing),
        "d2_missing_answer_cases": d2_missing,
        "d3_candidate_count": len(d3_candidates),
        "d3_candidates": d3_candidates,
        "d3_missing_reference_answer_count": len(d3_missing_reference),
        "d3_missing_reference_answer_cases": d3_missing_reference,
        "unknown_skill_count": len(unknown_skills),
        "unknown_skills": unknown_skills,
        "skill_counts": dict(skill_counter),
        "attachment_errors": attachment_errors,
        "ignored_attachment_rule": "files named ~$* or .DS_Store are ignored",
    }


def _portable_manifest(manifest: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{key: value for key, value in entry.items() if key != "absolute_path"} for entry in manifest]


def _portable_cases(cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    portable_cases: list[dict[str, Any]] = []
    for case in cases:
        portable_case = dict(case)
        portable_attachments: list[dict[str, Any]] = []
        for attachment in case.get("attachments", []):
            portable_attachments.append({key: value for key, value in attachment.items() if key != "original_path"})
        portable_case["attachments"] = portable_attachments
        portable_cases.append(portable_case)
    return portable_cases


def preprocess_dataset(
    excel_path: Path,
    attachments_root: Path,
    output_dir: Path,
    *,
    skills_workbook: Path | None = None,
    sessions_json: Path | None = None,
    repo_root: Path | None = None,
    portable_paths: bool = False,
) -> dict[str, Any]:
    repo_root = repo_root or Path.cwd()
    cases, manifest, attachment_errors = load_cases_from_excel(excel_path, attachments_root, repo_root=repo_root)
    cases_to_write = _portable_cases(cases) if portable_paths else cases
    manifest_to_write = _portable_manifest(manifest) if portable_paths else manifest
    domain_skills = load_domain_skill_registry(skills_workbook) if skills_workbook else {}
    runtime_skills = load_runtime_skill_registry(sessions_json) if sessions_json else {}
    report = build_quality_report(
        cases,
        manifest,
        attachment_errors,
        domain_skills=domain_skills,
        runtime_skills=runtime_skills,
    )
    write_jsonl(output_dir / "cases.jsonl", cases_to_write)
    write_json(output_dir / "attachment_manifest.json", manifest_to_write)
    write_json(output_dir / "data_quality_report.json", report)
    write_csv(
        output_dir / "cases.csv",
        [
            {
                "case_id": case["case_id"],
                "source_row": case["source_row"],
                "skill_name": case["skill_name"],
                "d2_enabled": case["d2_enabled"],
                "has_expected_answer": case["expected_answer"] is not None,
                "has_target_state": case["target_state"] is not None,
                "has_gold_chain": case["gold_chain"] is not None,
                "d3_candidate": case.get("d3_candidate"),
                "d3_priority": case.get("d3_priority"),
                "attachment_count": len(case["attachments"]),
                "annotation_status": ";".join(case["annotation_status"]),
                "user_query": case["user_query"],
            }
            for case in cases
        ],
        [
            "case_id",
            "source_row",
            "skill_name",
            "d2_enabled",
            "has_expected_answer",
            "has_target_state",
            "has_gold_chain",
            "d3_candidate",
            "d3_priority",
            "attachment_count",
            "annotation_status",
            "user_query",
        ],
    )
    return {
        "cases": cases,
        "attachment_manifest": manifest,
        "quality_report": report,
        "output_dir": str(output_dir),
    }
